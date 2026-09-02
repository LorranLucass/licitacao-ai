# services/item_matcher.py

from pathlib import Path
import re
import unicodedata

from openpyxl import load_workbook


# ============================================================
# CONFIGURAÇÃO
# ============================================================

# Caminho absoluto baseado na localização deste arquivo, e não
# na pasta de onde o processo foi iniciado. Necessário porque
# em ambientes serverless (Vercel) o diretório de trabalho não
# é garantido ser a raiz do projeto.
PASTA_RAIZ = Path(__file__).resolve().parent.parent

ARQUIVO_PRODUTOS = PASTA_RAIZ / "dados" / "produtos.xlsx"
ABA_PRODUTOS = "Pelotão DELTA"

TIPOS_PERMITIDOS = {
    "MONITOR",
    "TV",
}


# ============================================================
# PALAVRAS QUE DEVEM SER BLOQUEADAS
# ============================================================
# Se o produto da tabela possuir uma dessas palavras,
# ele NÃO poderá ser usado como monitor ou TV.
#
# Isso evita, por exemplo:
# - COMPUTADOR COM MONITOR
# - DESKTOP
# - NOTEBOOK
# - MINI PC
# - ALL IN ONE
#
# mesmo que o nome contenha a palavra MONITOR.

TERMOS_BLOQUEADOS_COMPUTADOR = [
    "COMPUTADOR",
    "DESKTOP",
    "NOTEBOOK",
    "LAPTOP",
    "MINI PC",
    "MINIPC",
    "PC GAMER",
    "PC",
    "ALL IN ONE",
    "ALL-IN-ONE",
    "ALLINONE",
    "WORKSTATION",
    "CHROMEBOOK",
    "THIN CLIENT",
]


# ============================================================
# TAMANHOS DE REFERÊNCIA
# ============================================================

TAMANHOS_MONITOR = [
    15.0,
    19.5,
    21.5,
    24.0,
]

TAMANHOS_TV = [
    32.0,
    43.0,
    50.0,
    55.0,
    60.0,
    65.0,
    75.0,
]


# ============================================================
# NORMALIZAÇÃO DE TEXTO
# ============================================================

def normalizar_texto(valor):
    """
    Normaliza texto para comparação:
    maiúsculas, sem acentos, sem espaços nas pontas.

    Exemplo:
        "Televisão 50\" Smart"
        ->
        "TELEVISAO 50\" SMART"
    """

    if valor is None:
        return ""

    texto = str(valor).upper().strip()

    # Remove acentos preservando o restante do texto
    # (inclusive aspas e símbolos usados para polegadas).
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(
        caractere
        for caractere in texto
        if not unicodedata.combining(caractere)
    )

    return texto


# ============================================================
# NORMALIZAÇÃO DE TAMANHO
# ============================================================

def normalizar_tamanho(valor):
    """
    Converte tamanhos para número decimal.

    Exemplos:

        24
        24"
        24”
        24 POLEGADAS
        19,5"
        21.5"

    Retorna:

        24.0
        19.5
        21.5
    """

    if valor is None:
        return None

    texto = str(valor).upper().strip()

    texto = texto.replace(",", ".")

    resultado = re.search(
        r"(\d+(?:\.\d+)?)",
        texto,
    )

    if not resultado:
        return None

    try:
        return float(resultado.group(1))

    except (ValueError, TypeError):
        return None


# ============================================================
# EXTRAÇÃO DO TAMANHO
# ============================================================

def extrair_tamanho(texto):
    """
    Reconhece:

        15"
        19,5"
        21.5"
        24"
        32"
        43"
        50"
        55"
        60"
        65"
        75"

    Também reconhece:

        24 POLEGADAS
        24 POLEGADA
        24 POL
        24 POLEG.
    """

    if not texto:
        return None

    texto = normalizar_texto(texto)

    padroes = [

        # 24"
        r'(\d+(?:[.,]\d+)?)\s*(?:"|”|″)',

        # 24 POLEGADAS
        r"(\d+(?:[.,]\d+)?)\s*POLEGADAS?",

        # 24 POL
        r"(\d+(?:[.,]\d+)?)\s*POL\b",

        # 24 POLEG.
        r"(\d+(?:[.,]\d+)?)\s*POLEG\.",

    ]

    for padrao in padroes:

        resultado = re.search(
            padrao,
            texto,
        )

        if resultado:

            return normalizar_tamanho(
                resultado.group(1)
            )

    return None


# ============================================================
# BLOQUEIO DE COMPUTADORES
# ============================================================

def produto_bloqueado(texto):
    """
    Verifica se o produto é um computador,
    notebook, desktop etc.

    Retorna:

        True  = produto bloqueado
        False = produto permitido
    """

    texto = normalizar_texto(texto)

    if not texto:
        return False

    for termo in TERMOS_BLOQUEADOS_COMPUTADOR:

        if termo == "PC":
            # PC precisa ser palavra isolada.
            if re.search(r"\bPC\b", texto):
                return True

        else:

            if termo in texto:
                return True

    return False


# ============================================================
# IDENTIFICAÇÃO DO TIPO
# ============================================================

def identificar_tipo(texto):
    """
    Identifica somente:

        MONITOR
        TV

    IMPORTANTE:

    Antes de identificar o tipo, verifica se o produto
    é um computador/notebook/desktop.

    Isso impede que um produto como:

        COMPUTADOR DESKTOP COM MONITOR 24"

    seja classificado como MONITOR.
    """

    if not texto:
        return None

    texto = normalizar_texto(texto)

    # ========================================================
    # BLOQUEIO ABSOLUTO DE COMPUTADOR
    # ========================================================

    if produto_bloqueado(texto):
        return None

    # ========================================================
    # MONITOR
    # ========================================================
    # \bMONITOR\b sozinho não reconhece o plural
    # "MONITORES". (MONITOR)(ES)? cobre singular e plural
    # sem confundir com "MONITORAMENTO".

    if re.search(r"\bMONITOR(ES)?\b", texto):

        return "MONITOR"

    # ========================================================
    # TV
    # ========================================================
    # Aceita as variações mais comuns usadas em editais:
    # TV, TVS, TELEVISOR(ES), TELEVISAO/TELEVISOES (o texto
    # já chega sem acento, ver normalizar_texto), SMART TV,
    # TV LED, LED TV etc. — todas contêm um destes termos.

    if "SMART TV" in texto:
        return "TV"

    if "TELEVISOR" in texto:
        return "TV"

    if "TELEVISAO" in texto or "TELEVISOES" in texto:
        return "TV"

    if re.search(r"\bTVS?\b", texto):
        return "TV"

    return None


# ============================================================
# VALIDAÇÃO DO TAMANHO
# ============================================================

def tamanho_permitido(tipo, tamanho):
    """
    Verifica se o tamanho está dentro da faixa trabalhada.

    MONITOR:
        15 até 24

    TV:
        32 até 75
    """

    if tamanho is None:
        return False

    if tipo == "MONITOR":

        return (
            min(TAMANHOS_MONITOR)
            <= tamanho
            <= max(TAMANHOS_MONITOR)
        )

    if tipo == "TV":

        return (
            min(TAMANHOS_TV)
            <= tamanho
            <= max(TAMANHOS_TV)
        )

    return False


# ============================================================
# EXTRAÇÃO DO MODELO
# ============================================================

def extrair_modelo(produto):
    """
    Extrai o modelo final do nome do produto.

    Exemplo:

        MONITOR 24" GAMER LED FULL HD HQ - 24HQ-LED

    Resultado:

        HQ - 24HQ-LED
    """

    if not produto:
        return ""

    texto = str(produto).strip()

    resultado = re.search(
        r"(HQ\s*-\s*.+)$",
        texto,
        re.IGNORECASE,
    )

    if resultado:

        return resultado.group(1).strip()

    return texto


# ============================================================
# CARACTERÍSTICAS COMPATÍVEIS
# ============================================================

def caracteristicas_compativeis(
    descricao_edital,
    produto,
):
    """
    Compara características importantes do edital
    com o produto da Tabela DELTA.
    """

    edital = normalizar_texto(
        descricao_edital
    )

    produto_nome = normalizar_texto(
        produto.get("produto", "")
    )

    # ========================================================
    # BLOQUEIO DE SEGURANÇA
    # ========================================================
    # Mesmo que alguma alteração futura permita MONITOR
    # durante a identificação, um computador nunca passa.

    if produto_bloqueado(produto_nome):
        return False

    # ========================================================
    # CÂMERA / WEBCAM
    # ========================================================

    exige_camera = any(
        termo in edital
        for termo in [
            "COM CÂMERA",
            "COM CAMERA",
            "WEBCAM",
            "CÂMERA INTEGRADA",
            "CAMERA INTEGRADA",
            "VIDEOCONFERÊNCIA",
            "VIDEOCONFERENCIA",
        ]
    )

    if exige_camera:

        possui_camera = any(
            termo in produto_nome
            for termo in [
                "CÂMERA",
                "CAMERA",
                "WEBCAM",
            ]
        )

        if not possui_camera:
            return False

    # ========================================================
    # 4K
    # ========================================================

    exige_4k = "4K" in edital

    if exige_4k and "4K" not in produto_nome:
        return False

    # ========================================================
    # QLED
    # ========================================================

    exige_qled = "QLED" in edital

    if exige_qled and "QLED" not in produto_nome:
        return False

    # ========================================================
    # OLED
    # ========================================================

    exige_oled = "OLED" in edital

    if exige_oled and "OLED" not in produto_nome:
        return False

    return True


# ============================================================
# PONTUAÇÃO POR PALAVRAS-CHAVE
# ============================================================
# Usada apenas para DESEMPATAR candidatos que já passaram
# pelos critérios obrigatórios (tipo + tamanho + segurança).
# Nunca decide sozinha se um produto é compatível.

PALAVRAS_CHAVE_BONUS = [
    "SMART",
    "LED",
    "4K",
    "QLED",
    "OLED",
    "FULL HD",
    "HD",
    "WI-FI",
    "WIFI",
    "USB",
    "HDMI",
    "IPS",
    "CURVO",
    "GAMER",
    "CONVERSOR DIGITAL",
    "CONTROLE REMOTO",
    "ANDROID",
]


def pontuacao_palavras_chave(descricao_edital, produto_nome):
    """
    Conta quantas palavras-chave relevantes aparecem em
    ambos os textos. Quanto maior, mais parecido.
    """

    edital = normalizar_texto(descricao_edital)
    produto = normalizar_texto(produto_nome)

    pontos = 0

    for palavra in PALAVRAS_CHAVE_BONUS:

        if palavra in edital and palavra in produto:
            pontos += 1

    return pontos


# ============================================================
# CARREGAR PRODUTOS DA TABELA DELTA
# ============================================================

def carregar_produtos(
    caminho=ARQUIVO_PRODUTOS,
):
    """
    Carrega somente TV e MONITOR da aba Pelotão DELTA.

    Mantém:

        marca
        produto
        modelo
        fabricante
        custo
        minimo_feirao
        tipo
        tamanho
    """

    arquivo = Path(caminho)

    if not arquivo.exists():

        raise FileNotFoundError(
            f"Tabela de produtos não encontrada: {arquivo}"
        )

    workbook = load_workbook(
        arquivo,
        data_only=True,
    )

    if ABA_PRODUTOS not in workbook.sheetnames:

        raise ValueError(
            f"A aba '{ABA_PRODUTOS}' não foi encontrada."
        )

    planilha = workbook[
        ABA_PRODUTOS
    ]

    produtos = []

    marca_atual = ""

    # ========================================================
    # PERCORRER PLANILHA
    # ========================================================

    for linha in planilha.iter_rows(
        min_row=2,
        values_only=True,
    ):

        # A estrutura principal da DELTA é:
        #
        # MARCA
        # PRODUTO
        # CUSTO
        # MIN. FEIRÃO

        marca = linha[0] if len(linha) > 0 else None
        produto = linha[1] if len(linha) > 1 else None
        custo = linha[2] if len(linha) > 2 else None
        minimo_feirao = (
            linha[3]
            if len(linha) > 3
            else None
        )

        # ====================================================
        # LINHA SEM PRODUTO
        # ====================================================

        if not produto:
            continue

        produto = str(
            produto
        ).strip()

        if not produto:
            continue

        # ====================================================
        # MARCA
        # ====================================================

        if marca:

            marca_atual = str(
                marca
            ).strip()

        # ====================================================
        # BLOQUEIO DE COMPUTADOR
        # ====================================================

        if produto_bloqueado(produto):

            continue

        # ====================================================
        # IDENTIFICAR TIPO
        # ====================================================

        tipo = identificar_tipo(
            produto
        )

        # Somente TV e MONITOR
        if tipo not in TIPOS_PERMITIDOS:
            continue

        # ====================================================
        # TAMANHO
        # ====================================================

        tamanho = extrair_tamanho(
            produto
        )

        # Produto sem tamanho não serve
        if tamanho is None:
            continue

        # Fora da faixa trabalhada
        if not tamanho_permitido(
            tipo,
            tamanho,
        ):
            continue

        # ====================================================
        # MODELO
        # ====================================================

        modelo = extrair_modelo(
            produto
        )

        # ====================================================
        # PRODUTO FINAL
        # ====================================================

        produtos.append(
            {
                # Mantém HQ como marca utilizada no bloco
                "marca": "HQ",

                # Nome completo da tabela DELTA
                "produto": produto,

                # Modelo extraído
                "modelo": modelo,

                # Fabricante
                "fabricante": "BELMICRO",

                # PREÇO DE CUSTO
                "custo": custo,

                # PREÇO MÍNIMO DO FEIRÃO
                "minimo_feirao": minimo_feirao,

                # Tipo
                "tipo": tipo,

                # Tamanho
                "tamanho": tamanho,
            }
        )

    return produtos


# ============================================================
# ENCONTRAR PRODUTO
# ============================================================

def encontrar_produto(
    descricao_edital,
    produtos,
):
    """
    Procura produto compatível na Tabela DELTA.

    Critérios obrigatórios:

        1. Tipo
        2. Tamanho
        3. Características
        4. Produto não pode ser computador

    O modelo e os preços vêm exclusivamente
    da Tabela DELTA.
    """

    # ========================================================
    # TIPO DO EDITAL
    # ========================================================

    tipo_edital = identificar_tipo(
        descricao_edital
    )

    if tipo_edital not in TIPOS_PERMITIDOS:

        return None

    # ========================================================
    # TAMANHO DO EDITAL
    # ========================================================

    tamanho_edital = extrair_tamanho(
        descricao_edital
    )

    if tamanho_edital is None:

        return None

    # ========================================================
    # TAMANHO DENTRO DA FAIXA
    # ========================================================

    if not tamanho_permitido(
        tipo_edital,
        tamanho_edital,
    ):

        return None

    # ========================================================
    # CANDIDATOS
    # ========================================================

    candidatos = []

    for produto in produtos:

        # ====================================================
        # SEGURANÇA: BLOQUEAR COMPUTADOR
        # ====================================================

        if produto_bloqueado(
            produto.get("produto", "")
        ):

            continue

        # ====================================================
        # TIPO
        # ====================================================

        if produto.get("tipo") != tipo_edital:

            continue

        # ====================================================
        # TAMANHO
        # ====================================================

        tamanho_produto = produto.get(
            "tamanho"
        )

        if tamanho_produto is None:

            tamanho_produto = extrair_tamanho(
                produto.get("produto", "")
            )

        if tamanho_produto is None:

            continue

        # ====================================================
        # TAMANHO EXATO
        # ====================================================

        if abs(
            tamanho_produto
            - tamanho_edital
        ) > 0.01:

            continue

        # ====================================================
        # CARACTERÍSTICAS
        # ====================================================

        if not caracteristicas_compativeis(
            descricao_edital,
            produto,
        ):

            continue

        candidatos.append(
            produto
        )

    # ========================================================
    # NENHUM PRODUTO
    # ========================================================

    if not candidatos:

        return None

    # ========================================================
    # ESCOLHER O CANDIDATO MAIS PARECIDO
    # ========================================================
    # Entre os candidatos que já passaram por tipo, tamanho e
    # características obrigatórias, escolhe o que tem mais
    # palavras-chave em comum com a descrição do edital
    # (ex: SMART, 4K, WI-FI). Em empate, mantém a ordem
    # original da planilha.

    candidatos_ordenados = sorted(
        candidatos,
        key=lambda produto: pontuacao_palavras_chave(
            descricao_edital,
            produto.get("produto", ""),
        ),
        reverse=True,
    )

    resultado = candidatos_ordenados[0].copy()

    # ========================================================
    # DADOS FIXOS DA EMPRESA
    # ========================================================

    resultado["marca"] = "HQ"

    resultado["fabricante"] = "BELMICRO"

    # ========================================================
    # MODELO VEM DA DELTA
    # ========================================================

    resultado["modelo"] = extrair_modelo(
        resultado.get("produto", "")
    )

    # ========================================================
    # TAMANHO
    # ========================================================

    resultado["tamanho"] = tamanho_edital

    return resultado


# ============================================================
# FILTRAR ITENS DO EDITAL
# ============================================================

def filtrar_itens(dados):
    """
    Filtra os itens do edital usando a Tabela DELTA.

    Somente itens que possuam produto compatível
    são mantidos.

    Os preços da DELTA são preservados.
    """

    produtos = carregar_produtos()

    itens_encontrados = []

    # ========================================================
    # PERCORRER ITENS DO EDITAL
    # ========================================================

    for item in dados.get(
        "itens",
        [],
    ):

        descricao = item.get(
            "descricao",
            "",
        )

        # ====================================================
        # IDENTIFICAR TIPO
        # ====================================================

        tipo = identificar_tipo(
            descricao
        )

        # Ignora tudo que não seja TV ou monitor
        if tipo not in TIPOS_PERMITIDOS:

            continue

        # ====================================================
        # ENCONTRAR PRODUTO
        # ====================================================

        produto = encontrar_produto(
            descricao,
            produtos,
        )

        # Se não encontrou produto compatível,
        # não adiciona o item.
        if not produto:

            continue

        # ====================================================
        # MONTAR ITEM FINAL
        # ====================================================

        item_filtrado = item.copy()

        # ====================================================
        # PRODUTO COMPLETO DA DELTA
        # ====================================================

        item_filtrado[
            "produto_tabela"
        ] = produto.get(
            "produto",
            "",
        )

        # ====================================================
        # MARCA
        # ====================================================

        item_filtrado[
            "marca_tabela"
        ] = produto.get(
            "marca",
            "HQ",
        )

        # ====================================================
        # MODELO
        # ====================================================

        item_filtrado[
            "modelo"
        ] = produto.get(
            "modelo",
            "",
        )

        # Compatibilidade com o bloco de notas
        item_filtrado[
            "modelo_tabela"
        ] = produto.get(
            "modelo",
            "",
        )

        # ====================================================
        # FABRICANTE
        # ====================================================

        item_filtrado[
            "fabricante"
        ] = produto.get(
            "fabricante",
            "BELMICRO",
        )

        # ====================================================
        # CUSTO
        # ====================================================
        # IMPORTANTE:
        # Não apagar o preço vindo da DELTA.

        item_filtrado[
            "custo"
        ] = produto.get(
            "custo",
            0,
        )

        # ====================================================
        # MÍNIMO FEIRÃO
        # ====================================================
        # IMPORTANTE:
        # Não apagar o preço vindo da DELTA.

        item_filtrado[
            "minimo_feirao"
        ] = produto.get(
            "minimo_feirao",
            0,
        )

        # ====================================================
        # TAMANHO
        # ====================================================

        item_filtrado[
            "tamanho"
        ] = produto.get(
            "tamanho",
            "",
        )

        # ====================================================
        # TIPO
        # ====================================================

        item_filtrado[
            "tipo"
        ] = produto.get(
            "tipo",
            tipo,
        )

        # ====================================================
        # ADICIONAR
        # ====================================================

        itens_encontrados.append(
            item_filtrado
        )

    # ========================================================
    # RETORNO
    # ========================================================

    dados_filtrados = dados.copy()

    dados_filtrados[
        "itens"
    ] = itens_encontrados

    return dados_filtrados